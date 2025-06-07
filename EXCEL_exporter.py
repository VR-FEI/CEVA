import os
import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import pytz

class ExcelExporter:
    def __init__(self):
        # Setup paths
        results_folder_path = os.path.join(os.path.dirname(os.path.realpath(__file__)), "SKU Results")
        self.temp_txt_file = os.path.join(results_folder_path, "temp.txt")
        
        # Create results folder if it doesn't exist
        if not os.path.exists(results_folder_path):
            os.makedirs(results_folder_path)
            
        # Generate timestamped Excel filename
        tz = pytz.timezone("America/Sao_Paulo")
        timestamp = datetime.datetime.now(tz).strftime("%Y-%m-%d_%H-%M-%S")
        self.excel_file = os.path.join(results_folder_path, f"SKU_Export_{timestamp}.xlsx")

    def read_txt_data(self):
        """Read data from temp.txt file"""
        if not os.path.exists(self.temp_txt_file):
            print(f"[ERRO] Arquivo temp.txt não encontrado: {self.temp_txt_file}")
            return []
            
        data = []
        try:
            with open(self.temp_txt_file, 'r') as file:
                for line in file:
                    # Split by comma and strip whitespace
                    parts = [part.strip() for part in line.strip().split(',')]
                    if len(parts) >= 3:  # Ensure we have at least SKU, description, family
                        data.append({
                            'sku': parts[0],
                            'description': parts[1],
                            'family': parts[2]
                        })
            return data
        except Exception as e:
            print(f"[ERRO] Falha ao ler arquivo temp.txt: {e}")
            return []

    def export_to_excel(self):
        """Export data from temp.txt to Excel file"""
        # Read data from temp.txt
        data = self.read_txt_data()
        if not data:
            print("[INFO] Nenhum dado para exportar.")
            return False

        # Create a new workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "SKU Data"

        # Add headers
        headers = ["SKU", "Description", "Family"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

        # Add data
        for row_num, item in enumerate(data, 2):
            ws.cell(row=row_num, column=1).value = item['sku']
            ws.cell(row=row_num, column=2).value = item['description']
            ws.cell(row=row_num, column=3).value = item['family']

        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width

        # Save the workbook
        try:
            wb.save(self.excel_file)
            print(f"[INFO] Dados exportados com sucesso para: {self.excel_file}")
            return True
        except Exception as e:
            print(f"[ERRO] Falha ao salvar arquivo Excel: {e}")
            return False
            
    def clear_temp_file(self):
        """Clear the temp.txt file after successful export"""
        try:
            open(self.temp_txt_file, 'w').close()
            print("[INFO] Arquivo temp.txt limpo após exportação.")
        except Exception as e:
            print(f"[ERRO] Falha ao limpar arquivo temp.txt: {e}")
            
    def export_skus(self):
        """Main method to export SKUs from txt to Excel"""
        success = self.export_to_excel()
        if success:
            self.clear_temp_file()
        return success